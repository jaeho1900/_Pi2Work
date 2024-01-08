# =====================
# Function
# =====================

# # 함수는 코드에 대한 이름표
# ┌ '함수명'으로 바로 사용: 내장함수, 사용자 정의 함수
# └ '객체명.함수명'으로 사용: 다른 모듈(패키지) 함수, 클래스 함수
# 내장 함수: print, open, int 등 기본 함수
# 사용자 정의 함수: def 를 활용하여 직접 정의한 함수
# 다른 모듈(패키지) 함수: 별도 파일에서 정의한 함수나 패키지의 함수('import 폴더명.파일명.함수명'으로 호출)
# 클래스 함수: 파이썬은 클래스로 모든 요소를 생성하며, 클래스의 객체는 '객체명.함수명'으로 사용할 수 있다
# 함수호출부: argument = 전달인자 = 값 = 인수
# 함수정의부: parameter = 매개변수 = 변수 = 인자


# =====================
# 내장 함수
# =====================

# >>> zip(): 동일 개수의 자료형을 묶어서 리스트로 반환(짝이 안 맞으면 맞는 만큼만 적용) -----

a = [1, 2, 3, 4, 5]
b = ['a', 'b', 'c', 'd']
list(zip(a, b))

# >>> enumerate(): 리스트, 튜플, 문자열을 인자로 주면 인덱스와 값을 함께 반환 -----

lst = ['lion', 'tiger', 'bear', 'hippo']

for i, lstName in enumerate(lst):
    print(i, lstName)

# >>> map(): iterable 객체 요소를 하나씩 지정 함수로 보냄, 결과값은 담을 그릇(리스트나 튜플) 필요 -----
    
x = [1, 2, 3, 4, 5]
list(map(float, x))
tuple(map(lambda x: x**3, x))

# >>> filter(): 참인 결과 값만 리스트로 반환 -----

def odd(x):
    return x % 2 != 0

lst = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
list(filter(odd, lst))

# >>> eval(): 입력 받은 문자열을 함수나 클래스에서 동적 실행 -----

eval('3 + 4')  # 7

# >>> dir(): 해당 객체가 가지고 있는 변수와 함수를 반환 -----

# 객체의 속한 메서드는 객체 뒤에서 .으로 호출하여 사용
dir([1, 2, 3])
dir([1, 2, 3]).clear()

# >>> id(): 객체의 고유 주소값 반환 -----

id(100)

# >>> isinstance(인스턴스, 클래스명): 해당 클래스의 인스턴스 여부를 boolean 반환 -----

x = list([1, 2, 3, 4, 5])
type(x)
isinstance(x, list)

# >>> globals(), locals() -----

globals()['test'] = 100	                # {'test': 100} 글로벌 변수에 추가
print(test)

for i in range(1, 4):
    locals()['test' + str(i)] = i * 10  # {생략, 'test1': 10, 'test2': 20, 'test3': 30}

print(globals())
print(locals())


# =====================
# 사용자 정의 함수
# =====================

# --------------------
# 매개변수가 없는 사용자 정의 함수
# --------------------

def pair_gugudan():
    for x in range(2, 10, 2):
        for y in range(1, 10):
            print(x, 'x', y, '=', x * y, end='\t')
        print(end='\n')

pair_gugudan()


# --------------------
# 매개변수가 있는 사용자 정의 함수
# --------------------

def show_list_element(ani):
    for i in ani:
        print(i)

animals = ['Tiger', 'Wolf', 'Lion']
show_list_element(animals)


# >>> 가변 매개변수 -----

# 가변 매개변수는 매개변수 앞에 *를 붙여주어 튜플 객체로 저장
# 가변 매개변수는 하나만 사용할 수 있으며, 가변 매개변수 뒤에는 일반 매개변수가 올 수 없다.
def print_n_times(n, *values):
    for i in range(n):
        for i in values:
            print(i)
        print()

print_n_times(3, "안녕,", "즐거운", "파이썬 코딩이예요.")

# 호출할 때 가변 인수 활용(리스트/튜플)
def foo(a, b, c):
    print(a, b, c)

data = [1, 2, 3]
foo(*data)

# >>> 기본(지정) 매개변수 -----

# '매개변수 = 값'의 형태로 입력하며, 매개변수를 입력하지 않으면 기본(지정)값이 설정
# 기본 매개변수 뒤에는 일반 매개변수가 올 수 없다
def print_n_times(value, n=2):
    for i in range(n):
        print(value)

print_n_times("안녕")  # n이 입력되지 않았으므로 n=2

# >>> 가변 매개변수가 기본 매개변수보다 앞에 올 때는 기본 매개변수의 기본 값으로 동작 -----

# 기본 매개변수의 기본 값을 변경하고 싶으면, 기본 매개변수명을 사용해야만 함
def print_n_times(*values, n=2):
    for i in range(n):
        print(values)
        print()

print_n_times("안녕,", "즐거운", "파이썬", 3)    # n=2 로 작동됨(3이 아님)
print_n_times("안녕,", "즐거운", "파이썬", n=3)  # n=3

# >>> 키워드 매개변수: 매개변수명을 지정해서 호출 -----

# 함수 파라미터 앞에 **를 붙여주며, 키워드와 값이 딕셔너리객체로 저장됨
def foo(**kwargs):
    print(kwargs)

foo(a=1, b=2, c=3)

# >>> 가변매개변수와 키워드매개변수를 함께 사용하면 결과는 튜플과 딕셔너리로 분리됨 -----

def foo(*args, **kwargs):
    print(args)       # 튜플객체 반환
    print(kwargs)     # 딕셔너리객체 반환

foo(1, 2, 3, a=1, b=1, c=2)


# --------------------
# 리턴값이 있는 사용자 정의 함수
# --------------------

# >>> 리턴은 결과값을 반환하며 함수를 중단하는 역할 -----

def f(treeHit):  # treeHit: 파라미터, 매개변수, 인자
    while treeHit < 10:
        treeHit = treeHit + 1
        print("나무를 %d번 찍었습니다." % treeHit)
        if treeHit == 5:
            return "나무 넘어갑니다."
    return -1

print(f(3))     # 3: 인수, 전달인자


# --------------------
# yield Method (제너레이터 함수)
# --------------------

# return 이 만든 빵을 전달한 후 다시 빵을 만들 때 다시 옷을 갈아입고 손을 씻는다면,
# yield 는 만든 빵을 전달한 후(코드중지상태를 저장) 즉시 빵을 만드는 것

# yield 가 사용된 함수는 ()를 붙여도 코드가 바로 실행하지 않고, 객체만 생성
# next() 호출을 통해 제너레이터 객체 내의 코드를 실행하여 값을 리턴 후 현상태를 유지
# 다시 next() 호출을 하면 중단된 시점에서 코드가 재실행
def product(n):
    for i in range(n):
        num = "Num." + str(i)
        yield num

def packed(num):
    print("{}, packaging completed".format(num))

for i in product(10):
    packed(i)

# >>> 리스트 요소를 하나씩 전달 -----
    
def number_generator():
    x = [1, 2, 3]
    yield from x

k = number_generator()
num1 = next(k)
num2 = next(k)
print(num1, num2)


# =====================
# 모듈(패키지) 함수
# =====================

# >>> 진행율 그래프 -----

from tqdm import tqdm
for i in tqdm(range(1, 10000000, 1)):
    i += i
print(i)

# >>> os() -----

import os
os.getcwd()                              # 현재 경로
__file__                                 # 현재 파일의 경로(파일명 포함)
ds_path = os.getcwd()
ds_file_path = __file__
os.path.abspath(ds_file_path)            # 절대경로로 변환
os.path.dirname(ds_file_path)            # 마지막 경로,파일명을 제외하고 반환
os.path.basename(ds_file_path)           # 마지막 경로,파일명을 반환
os.path.split(ds_file_path)              # 마지막 경로,파일명을 분리 반환
os.path.exists(ds_file_path)             # 파일 또는 경로가 존재 여부 반환
os.path.join(ds_path + os.sep + 'Test')  # 경로 병합
dir_name = os.path.join(ds_path + os.sep + 'Test')
os.mkdir(dir_name)                       # 새로운 폴더를 만들기
os.chdir(dir_name)                       # 작업 경로 변경
os.listdir(dir_name)                     # 경로 안의 하위경로명 및 파일명 리스트
os.rename('test0.txt', 'test6.txt')      # 파일명 또는 경로명 바꾸기
os.remove('test1.txt')                   # 파일삭제
os.system('copy test3.txt test7.txt')    # 시스템의 유틸리티나 dos명령어 사용
os.system('del *.txt')
os.chdir(os.path.dirname(dir_name))
print(os.getcwd())
os.rmdir(dir_name)                       # 폴더 삭제

# >>> datetime() -----

import datetime
print(datetime.datetime.__doc__)

# 날짜와 시간 출력
current = datetime.datetime(2020, 5, 16, 23, 10, 1, 100)
# 마이크로세컨드 제외
current_time = current.replace(microsecond=0)
# 날짜 출력
currentdate = datetime.datetime.now().date()
currentdate2 = datetime.date.today()
# 시간 출력
currenttime = datetime.datetime.now().time()
# 날짜와 시간 합치기
combine_datetime = datetime.datetime.combine(currentdate, currenttime)
# timestamp 는 float형!!
timestamp = datetime.datetime.now().timestamp()
# timestamp를 datetime형으로 변환
fts = datetime.datetime.fromtimestamp(timestamp)
# datetime을 원하는 형식의 str로 변환
chg = datetime.datetime.strftime(current, '%b %d %y %H %M %S')
chg2 = current.strftime('%b %d %y %H %M %S')
# str을 datetime으로 변환
todatetime = datetime.datetime.strptime(chg2, '%b %d %y %H %M %S')
# 날짜 차이
days_cap = (combine_datetime - current).days
# 시간 차이
hours_cap = (combine_datetime - current).seconds / 3600
# 분 차이
mins_cap = (combine_datetime - current).seconds / 60
# 초 차이
scds_cap = (combine_datetime - current).seconds
# 날짜계산과 시간계산을 위해서는 timedelta클래스를 사용!!
base_day = datetime.timedelta(days=3)
# 이전날짜
previous_day = current - base_day
# 다음날짜
next_day = current + base_day


# =====================
# Class 함수
# =====================

# >>> Man 이라는 객체를 만들기 위한 클래스 함수 정의: 클래스 공간에 데이터 저장 -----

class Man:
    cnt = 0                       # 전역변수: "클래스명.변수명"으로 사용, 클래스공간에 {"cnt":0} 를 저장
    class_var = 1000
    class_list_var = [2000]

    def __init__(self, name):     # 생성자: 객체생성될때 자동실행(객체 초기화에 활용)
        Man.cnt += 1
        self.name = name          # 지역변수: "self.변수명"으로 사용
        self.init_var = 3000      # 객체공간에 변수 생성: 점(.)의 의미는 p 객체공간을 의미하며, {"init_var":3000} 를 저장
        self.init_list_var = [4000]
        print('({}이 생성되었습니다.)'.format(self.name))

    def die(self):
        Man.cnt -= 1
        print('{}는 죽었습니다.!!!'.format(self.name))
        if Man.cnt == 0:
            print('{}는 최후의 생존자였습니다'.format(self.name))
        else:
            print('아직 {:d}명의 생존자가 남았습니다'.format(Man.cnt))

    @classmethod
    # class의 메소드(함수)는 인스턴스만 호출 할 수 있으나,
    # @classmethod를 사용하면 자기 메소드를 호출 할 수 있다.
    def how_many(how):
        print('현재 {}명이 남았습니다'.format(Man.cnt))

# 인스턴스(실체화된 객체) 생성, 클래스와 객체는 각각의 메모리 공간을 갖는다
Actor1 = Man('맨1')
Actor2 = Man('맨2')
# 인스턴스는 메소드를 호출한다
Actor1.die()
Actor2.die()
Actor3 = Man('맨3')
# 클래스는 메소드를 호출 할 수 없다(TypeError)
Man.die()
# 클래스는 클래스 전역변수는 호출 할 수 있다
Man.cnt
Man.class_var
# 클래스는 @classmethod로 정의된 메소드는 호출 할 수 있다
Man.how_many()
# 인스턴스는 변수를 호출한다
print(f'유닛이름 : {Actor1.name}, 값 : {Actor1.class_var}, {Actor1.init_var}')
# 인스턴스는 각자의 메모리 영역에 클래스변수를 저장
print(Actor1.class_var, Actor2.class_var)
Actor1.class_var = 5000
print(Actor1.class_var, Actor2.class_var)
# [예외] "리스트" 클래스변수는 같은 메모리 영역을 공유
print(Actor1.class_list_var[0], Actor2.class_list_var[0])
Actor1.class_list_var[0] = 6000
print(Actor1.class_list_var[0], Actor2.class_list_var[0])

# >>> @property를 붙여주면 속성처럼 사용 가능(호출을 위한 괄호()가 불필요) -----

class Car:

    def __init__(self, model):
        self.model = model

    @property
    def get_model(self):
        return self.model

c = Car("GV80")
print(c.get_model)  # c.get_model()이 아님
print(c.model)

# >>> 삼각김밥 클래스 -----

class Samgak:
    def __init__(self):
        self.source = '기본소스'
        self.kim = '광천김'
        self.bab = '쌀밥'
        self.food = '야채'

    def set_source(self, source_name):
        self.source = source_name

    def change_kim(self, kim_name):
        self.kim = kim_name

    def change_bab(self, bab_name):
        self.bab = bab_name

    def set_food(self, food_name):
        self.food = food_name

    def print(self):
        s1 = 'BlockDMask 가 맛있는 삼각김밥을 만들었습니다.\n'
        s1 += f'김은 {self.kim} 입니다.\n'
        s1 += f'밥은 {self.bab} 사용 하였고\n'
        s1 += f'소스는 {self.source} 촵촵 뿌리고\n'
        s1 += f'메인은 {self.food}을 넣었습니다.\n'
        print(s1)

# 참치 삼각김밥 인스턴스
chamchi = Samgak()
chamchi.print()
chamchi.set_food('동원참치')
chamchi.set_source('마요네즈')
chamchi.print()

chamchi.source
del chamchi.source
del chamchi

# >>> inheritance(상속) -----

# 기존 클래스가 라이브러리 형태로 제공되거나 수정이 허용되지 않는 상황일 때,
# 상속을 사용해서 기존 클래스를 변경하지 않고 기능을 추가하거나 변경한다
class Animal():         # 부모 클래스

    def walk(self):
        print('걷는다')

    def eat(self):
        print('먹는다')

    def greet(self):
        print('인사한다')

class Human(Animal):    # 자식 클래스
    def wave(self):
        print('손을 흔든다')

    def greet(self):    # 자식 클래스의 greet()가 우선
        self.wave()
        super().greet()  # super(): 부모 클래스의 greet()를 호출

person = Human()
person.walk()
person.eat()
person.wave()
person.greet()


# --------------------
# 예외(에러) 처리
# --------------------

# >>> 기본 구문 -----

classrooms = {'class1': [172, 188, 189, 159], 'class2': [191, 180]}

try:
    for class_id, heights in classrooms.items():
        for height in heights:
            if height > 190:
                print(class_id, '에 190이 넘는 학생이 있습니다')
                # 강제로 에러발생시켜서 종료처리
                raise StopIteration
# 정의된 에러
except StopIteration:
    pass
# 미정의된 에러들
else:
    print('알 수 없는 오류가 발생 되었습니다')
# 무조건 수행(file.close()등)
finally:
    print('예외 발생해도 무조건 수행 됩니다')

# >>> Exception은 모든 종류의 에러메세지를 반환 -----
    
try:
    list = []
    print(list[0])
except Exception as err:
    print('오류: ', err)

# >>> traceback 모듈은 에러메세지와 에러발생위치를 함께 반환 -----
    
try:
    list = []
    print(list[0])
except Exception:
    import traceback
    traceback.print_exc()

# >>> 사용자 정의 예외(에러): 기본적으로 Exception 클래스를 상속 받아서 사용 -----

class MyError(Exception):
    """Base 에러 클래스"""
    pass

class InsufficientFundsAsk(MyError):
    def __str__(self):
        return "잔고가 부족합니다."


# =====================
# 조건문, 반복문
# =====================

# # if문: if:_[elif:]_[else:]

# # while문: while:_[break]_[continue]_[pass]_[else:]

# # for문 : for_in_:[else:]


# =====================
# 파일 입출력 및 출력물 저장
# =====================

# --------------------
# open()
# --------------------

# r  읽기
# w  쓰기: 파일 만들고, 기존 파일 있으면 덮어쓰기
# x  쓰기: 파일 만들고, 기존 파일 있으면 오류 발생
# a  붙여쓰기: 기존 파일 내용 뒤에 추가하고, 없으면 새로 만듦
# r+ 붙여쓰기: 기존 파일 내용 앞에 추가
# b  바이너리모드로 열기

# >>> 읽기 -----

filepath = './trainingdata/2018_president_message.txt'
f = open(filepath,
         'r',
         encoding='cp949',  # 'cp949' for Windows
         errors='ignore')   # 'ignore' 에러무시, 'backslashreplace' 미지원 문자를 escape sequence 변환
print(f.read())          # 파일내용 전체를 가져와 문자열로 반환
# print(f.readlines())   # 파일내용 전체를 가져와 리스트로 반환: 줄마다 리스트 원소가 됨
f.close()

# readline(): 파일의 한 줄을 가져오며, 파일 포인터는 다음줄로 이동
with open(filepath, 'r', encoding='cp949') as f:
    matrix_quotes_list = []
    line = f.readline()
    while line != '':
        matrix_quotes_list.append(line)
        line = f.readline()
print(matrix_quotes_list)

# >>> 쓰기 -----

data = '한줄\n두줄\n세줄\n'
f = open('test_open.txt', 'w', encoding='utf8')   # 'cp949' for Windows
f.writelines(data)  # 한 번에 쓰기(list유형)
f.close()

# write(): 한 줄씩 쓰기(for 문 활용)
with open('test_open.txt', 'a', encoding='utf8') as f:
    for line in data:
        f.write(line)

# >>> 특정 화면에 출력되는 결과를 파일로 저장 -----

import sys
f = open('output.txt', 'w')
print('파일로 저장됩니다', file=f)
f.close()

# >>> 모든 화면에 출력되는 결과를 파일로 저장 -----

import sys
temp = sys.stdout
with open("output.txt", "w") as sys.stdout:  # 이어쓰기 옵션 'a'
    for i in range(10):
        print(i)
    sys.stdout = temp

# >>> 명령 프롬프트에서 저장 -----

# python test.py > output.txt    # 이어쓰기 옵션: >>


# --------------------
# 데이터프레임을 이미지로 저장
# --------------------

import dataframe_image as dfi
dfi.export(df, '파일명.jpg', max_rows=-1)  # -1은 모든 행을 의미
dfi.export(df, '파일명.jpg', max_rows=15)  # 행 개수를 15개로 제한(줄임 표시)


# --------------------
# pickle()
# --------------------

# >>> 읽기 -----

import pickle
fp = open('pickle_dict.pickle', 'rb')  # 바이너리임을 알수있도록 r뒤에 b를 붙인다
load = pickle.load(fp)                 # 파일에서 객체를 불러오기
print(load)
fp.close()

# >>> 쓰기 -----

fp = open('pickle_dict.pickle', 'wb')  # 바이너리임을 알수있도록 w뒤에 b를 붙인다
dic = {'name': 'Park', 'age': 50, 'hobby': ['soccer', 'baseball', 'basket']}
pickle.dump(dic, fp)                   # 객체를 파일에 저장
fp.close()


# --------------------
# openpyxl()
# --------------------

# >>> 읽기 -----

import openpyxl
wb = openpyxl.load_workbook(r'.\trainingdata\excel_sample.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
# ws = wb.worksheets[0]

# Multi Worksheets 읽기
for i, sheetname in enumerate(wb.sheetnames):
    locals()["ws{}".format(i)] = wb[sheetname]

# >>> 쓰기 -----
    
wb.save('stats_save.xlsx')

# >>> 행열 갯수 정보 -----

ws.max_column
ws.max_row

# >>> 행열 삽입, 삭제 -----

ws.insert_rows(2)
ws.insert_cols(2)
ws.delete_rows(2)
ws.delete_cols(2)

# >>> 셀 접근 -----

# 셀 자체
ws['A']                         # A열의 모든 셀
ws['A1']
ws.cell(row=1, column=1)
# 셀 내용 확인
for cell in ws['A']:            # A열의 모든 셀을 확인
    print(cell.value)
ws['B1'].value
ws.cell(row=1, column=2).value  # 1부터 시작!!

# >>> 셀 입력 -----

ws['A1'] = "구분"
ws.cell(row=4, column=2, value='대국')
# 예시
title = ['이름', '팩스', '전화', '제목', '메모']
for kwd, j in zip(title, list(range(1, len(title) + 1))):
    ws.cell(row=1, column=j).value = kwd

# >>> 셀 삭제 -----
    
ws.delete_rows(29, 73)
ws.delete_cols(8, 10)

# >>> 셀 스타일 -----

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import Color, PatternFill

# 글꼴
font_15 = Font(name='맑은 고딕', size=15, bold=True)
# 정렬
align_center = Alignment(horizontal='center', vertical='center')
align_vcenter = Alignment(vertical='center')
# 테두리
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
# 셀 색상
fill_orange = PatternFill(patternType='solid', fgColor=Color('FFC000'))
# 스타일 적용
for row in ws['B2:J2']:
    for cell in row:
        cell.font = font_15
        cell.alignment = align_center
        cell.fill = fill_orange
        cell.number_format = '#,##0.00%'


# =====================
# 파이썬 지원 프로그램
# =====================

# --------------------
# pyinstaller
# --------------------

# dist폴더에 exe 파일 생성
# cmd > pyinstaller -F 파일명.py  # '하나의 파일'지정 옵션
# cmd > pyinstaller -w 파일명.py  # gui등 터미널없이 윈도우모드로 실행 옵션
# cmd > pyinstaller -w --add-data '.\gui_basic\*.png;gui_basic'
#       --add-data '.\gui_basic\*2.png;gui_basic2' -F -i 'test.ico' 파일명.py
#       # 파일에 필요한 그림파일;폴더위치'로 추가 및 실행파일 icon 삽입 명령

# [option]
# --onefile        하나의 실행파일로 생성
# --noconsole      콘솔창을 안 띄우고 실행
# --icon=test.ico  실행파일의 아이콘 변경(https://icoconvert.com/)

# ※ ERROR : No module named 'pkg_resources.py2_warn'
# pip uninstall pyinstaller
# pip install https://github.com/pyinstaller/pyinstaller/archive/develop.zip


# --------------------
# VSCODE
# --------------------

# >>> VScode terminal 설정하기 - Windows -----

# * VScode terminal git bash로 설정 (웹개발 환경)
# - VScode에 들어가 ctrl + , 를 눌러, 설정에 들어간다

# - 검색창에 terminal.integrated.shell.windows 를 입력한다
# - Edit in settings.json 을 클릭한다
# - termianl.integrated.sheel.windows에 자신의 Git 설치 경로\bin\bash.exe\를 입력해준다.
# (예시)"terminal.integrated.shell.windows": "C:\\Program Files\\Git\\bin\\bash.exe"
# - 설정을 하고 난 후, vscode를 한 번 껐다가 켜준다
# - ctrl + shift + `를 눌러 새 터미널을 확인한다.

# * Terminal 변경하기 (Anaconda 환경)
# - Ctrl + Shift + P(팔레트 선택)을 누르고 명령어에 shell을 친다.
# - 아래의 Select Default Shell을 선택하면 shell을 선택할 수 있고,
# - Commnad Prompt를 선택하면 Anaconda Activation 환경을 사용할 수 있다.


# --------------------
# CONDA
# --------------------

# >>> 패키지 관리 -----

# 1. 아나콘다 환경 업데이트
# 1-1. 기본 뼈대 업데이트
# conda update -n base conda

# 1-2. 패키지 업데이트
# conda update --all

# 1-3. 패키지 관리 시스템인 PIP 업데이트
# python -m pip install --upgrade pip

# 2. 파이썬 버전 변경
# 2-1. 현재 변경 가능한 파이썬 버전 검색
# conda search python

# 2-2-1. 기존 환경은 유지하고, 새로운 가상환경을 만들어서 원하는 파이썬 버전을 적용하는 법
# conda create -n py3_11 python=3.11.0 anaconda

# 2-2-2. 아나콘다의 기본 파이썬 버전을 아예 원하는 버전으로 바꾸는 명령 (권장하지 않음!)
# conda install python=3.11.0
# conda update --all

# 3. 기타명령
# conda --version
# python --version
# conda list
# conda config --show channels                # 채널 검색
# conda config --add channels conda-forge     # 채널 추가
# conda config --remove channels conda-forge  # 채널 삭제
# conda config --set ssl_verify False         # 통신 체크 끄기
# conda install -c conda-forge 패키지명
# conda uninstall python

# >>> 실행 환경 -----
        
# conda env list
# conda list -n <myenv>
# conda create -n <myenv> python=3.9
# conda create --name <myenv> --file environment.yml
# conda activate <myenv>
# conda deactivate
# conda remove -n <myenv> --all
# # 실행 환경 파일 저장
# conda env export > environment.yml
# # 실행 환경 파일 불러오기(주 디렉토리에 다운 로드 필요)
# conda env update -f environment.yml
        

# --------------------
# 쥬피터
# --------------------

# >>> Jupyter 사용법 -----

# View   > Toggle Line Numbers>Cell line number
# Kernel > restart & clear output(이전 결과 삭제), restart & Run all(이전 결과 재실행)

# ESC: 명령모드
# Enter 혹은 Cell 클릭: 입력모드

# 명령 모드 a: 위 Cell 추가
# 명령 모드 b: 아래 Cell 추가

# 명령 모드 x: 셀 오려두기
# 명령 모드 c: 셀 복사
# 명령 모드 v: 셀 붙여넣기

# 명령 모드 dd: 셀삭제
# 명령 모드 z: 셀삭제 취소

# 명령 모드 m: Cell을 Markdown으로 설정
# 명령 모드 y: Cell을 code로 설정

# 명령 모드 Shift + M: 아래 셀과 합침
# 입력 모드 Ctrl + shift + -: 셀 분할

# 입력 모드 Ctrl + ] or Ctrl + [: 들여쓰기
# 입력 모드 Shift + Tab: 툴팁 표시

# 입력 모드 Ctrl + Enter: 셀 실행
# 입력 모드 Shift + Enter: 셀 실행 후 다음 셀로 이동

# %reset > y > enter # 모든 변수 삭제

# >>> Jupyter 확장 프로그램 -----

# conda install jupyter_contrib_nbextensions
# http://localhost:8888/nbextensions 에서 선택/적용

# pip uninstall jupyter
# pip uninstall jupyter_contrib_nbextensions
# pip uninstall jupyter_nbextensions_configurator

# conda remove --force jupyter notebook
# conda remove --force jupyter_contrib_nbextensions
# conda remove --force jupyter_nbextensions_configurator

# >>> Jupyter 테마 설치 -----

# conda install jupyterthemes

# 명령라인 커맨드 옵션
# jt  [-h] [-l] [-t THEME] [-f MONOFONT] [-fs MONOSIZE] [-nf NBFONT]
#     [-nfs NBFONTSIZE] [-tf TCFONT] [-tfs TCFONTSIZE] [-dfs DFFONTSIZE]
#     [-m MARGINS] [-cursw CURSORWIDTH] [-cursc CURSORCOLOR] [-vim]
#     [-cellw CELLWIDTH] [-lineh LINEHEIGHT] [-altp] [-altmd] [-altout]
#     [-P] [-T] [-N] [-r] [-dfonts]

# jt -l          # 사용가능한 테마목록을 출력
# jt -r          # 디폴트 모드로 변경
# jt -t onedork  # 테마 적용
# jt -t onedork -f roboto -fs 12 -altp -tfs 12 -nfs 12 -cellw 80% -T -N
# jt -t onedork -fs 115 -nfs 125 -tfs 115 -dfs 115 -ofs 115 -cursc r -cellw 80% -lineh 115 -altmd -kl -T -N
# -fs 115 : 코드 폰트 사이즈
# -nfs 125 : 노트북 메뉴 폰트 사이즈
# -tfs 115 : 마크다운 폰트 사이즈
# -dfs 115 : pandas DataFrame 폰트 크기
# -ofs 115 : Output 영역 폰트 크기
# -cursc r : cursor 색 red (onedork theme에서는 red가 가장 눈에 띄는 듯)
# -cellw 80% : 셀 가로 길이 80% (숫자 클수록 화면에 꽉참)
# -lineh 115 : 코드 줄 간격
# -altmd : 마크다운 셀의 배경을 투명하게 하는 옵션
# -kl : 커널 로고 표시 (노트북 화면 우상단에 python 로고)
# -T : 메뉴탭 아래에 툴바 표시 (저장, 셀 추가/삭제/이동, 커널 중단/재실행 등)
# -N : 노트북 화면에서 파일명 표시

# >>> 노트북(ipynb)과 스크립트(py) 변환 -----

# [1] ipynb를 다른 형식으로 변환

# $ pip install nbconvert
# $ jupyter nbconvert --to script a.ipynb [a.py]
# $ jupyter nbconvert --to html a.ipynb [a.html]

# 여러 파일 변환법
# $ jupyter nbconvert --to script a.ipynb b.ipynb ...

# 특정 폴더 변환법
# $ jupyter nbconvert --to script *.ipynb

# 경로 입력시에는 큰따옴표 사용
# $ jupyter nbconvert --to script "C:\*.ipynb"

# [2] ipynb와 py를 양방향으로 변환

# $ pip install ipynb-py-convert
# $ ipynb-py-convert a.py a.ipynb
# $ ipynb-py-convert a.ipynb a.py

# 스크립트를 노트북으로 바꿀때 셀구분이 하고 싶을 것이다.
# # % %(%%붙여쓴다) 로 셀구분이 가능하며, ''' 을 사용한 주석으로 마크다운을 적용할 수 있다.

# >>> 슬라이드 쇼 (인터넷 연결 필요) -----

# jupyter nbconvert Untitled.ipynb --to slides --post serve

# RISE: nbconvert 의 슬라이드 쇼 확장 프로그램 (주피터 노트북에서 바로 쇼 실행)
# conda install rise
# jupyter-nbextension enable rise --py --sys-prefix

# 사용
# Alt +R
